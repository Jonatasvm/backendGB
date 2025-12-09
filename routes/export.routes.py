from flask import Blueprint, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

export_bp = Blueprint('export', __name__)

@export_bp.route('/api/export/xls', methods=['POST'])
def export_xls():
    try:
        data = request.get_json()
        registros = data.get('registros', [])

        if not registros:
            return jsonify({'error': 'Nenhum registro selecionado'}), 400

        # Criar workbook e worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagamentos"

        # Definir cabeçalhos (ajuste conforme seus campos)
        headers = ['ID', 'Nome', 'Valor', 'Data', 'Status', 'Descrição']
        ws.append(headers)

        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Adicionar dados
        for registro in registros:
            row = [
                registro.get('id', ''),
                registro.get('nome', ''),
                registro.get('valor', ''),
                registro.get('data', ''),
                registro.get('status', ''),
                registro.get('descricao', '')
            ]
            ws.append(row)

        # Ajustar largura das colunas
        column_widths = [10, 30, 15, 15, 15, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Salvar em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Gerar nome do arquivo com data
        filename = f"pagamentos_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao gerar XLS: {e}")
        return jsonify({'error': 'Erro ao gerar arquivo'}), 500