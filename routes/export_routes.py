from flask import Blueprint, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime, timedelta
from db import get_connection

export_bp = Blueprint('export', __name__)

def normalize_forma_pagamento(forma_pagamento):
    """
    Normaliza a forma de pagamento para Título Case.
    PIX -> Pix, BOLETO -> Boleto, CHEQUE -> Cheque
    """
    if not forma_pagamento:
        return ''
    
    # Converter para maiúsculas e remover espaços
    forma_upper = str(forma_pagamento).strip().upper()
    
    if forma_upper == 'PIX':
        return 'Pix'
    elif forma_upper == 'BOLETO':
        return 'Boleto'
    elif forma_upper == 'CHEQUE':
        return 'Cheque'
    else:
        # Se não reconhecer, retornar com primeira letra maiúscula
        return str(forma_pagamento).strip().capitalize() if forma_pagamento else ''

def normalize_text_field(text):
    """
    Normaliza um texto para Título Case (primeira letra maiúscula).
    """
    if not text:
        return ''
    
    text = str(text).strip()
    return text[0].upper() + text[1:].lower() if len(text) > 0 else ''

@export_bp.route('/api/export/xls', methods=['POST'])
def export_xls():
    data = request.get_json()
    registros = data.get('registros', [])

    if not registros:
        return jsonify({'error': 'Nenhum registro selecionado'}), 400

    # Criar workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha de Importação"

    # Definir cabeçalhos (baseado nos campos do seu formulário)
    headers = [
        'ID',
        'Data Pagamento',
        'Valor',
        'Forma de Pagamento',
        'Quem Paga',
        'Centro de Custo',
        'Titular',
        'CPF/CNPJ',
        'Chave Pix',
        'Obra',
        'Categoria',
        'Status Lançamento',
        'Observação'
    ]
    ws.append(headers)

    # Estilizar cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Adicionar dados
    for registro in registros:
        # --- LOG PARA DEBUG ---
        print('DEBUG registro completo:', registro)
        print('DEBUG dataPagamento raw:', repr(registro.get('dataPagamento', '')))
        print('DEBUG valor raw:', repr(registro.get('valor', 0)))
        print('DEBUG id raw:', repr(registro.get('id', '')))
        # --- FIM LOG ---

        # Corrigir ID: remover aspas
        id_final = str(registro.get('id', '')).lstrip("'").strip()
        
        # Corrigir data: garantir datetime ou string sem aspa
        data_pagamento_raw = registro.get('dataPagamento', '')
        data_pagamento_final = ''
        if data_pagamento_raw:
            if isinstance(data_pagamento_raw, str):
                data_sem_aspa = data_pagamento_raw.lstrip("'").strip()
                try:
                    from datetime import datetime as dt
                    data_obj = dt.strptime(data_sem_aspa, '%Y-%m-%d')
                    data_pagamento_final = data_obj + timedelta(days=1)
                except Exception as e:
                    print(f'DEBUG erro ao parsear data: {e}')
                    # Se não for formato data, salva string sem aspa
                    data_pagamento_final = data_sem_aspa
            elif isinstance(data_pagamento_raw, (datetime,)):
                data_pagamento_final = data_pagamento_raw
            else:
                data_pagamento_final = str(data_pagamento_raw).lstrip("'").strip()
        else:
            data_pagamento_final = ''

        # Corrigir valor: garantir float, nunca string com aspa
        valor_raw = registro.get('valor', 0)
        if isinstance(valor_raw, str):
            valor_sem_aspa = valor_raw.lstrip("'").strip()
            try:
                valor_final = float(valor_sem_aspa) / 100
            except Exception as e:
                print(f'DEBUG erro ao converter valor: {e}')
                valor_final = 0.0
        else:
            try:
                valor_final = float(valor_raw) / 100 if valor_raw else 0.0
            except Exception as e:
                print(f'DEBUG erro ao converter valor: {e}')
                valor_final = 0.0
        
        print(f'DEBUG final - ID: {repr(id_final)}, DATA: {repr(data_pagamento_final)}, VALOR: {repr(valor_final)}')

        forma_pagamento = registro.get('formaDePagamento', '')
        forma_pagamento_normalizada = normalize_forma_pagamento(forma_pagamento)
        quem_paga_normalizado = 'Empresa'
        obra_raw = registro.get('obra', '')
        obra_normalizada = normalize_text_field(str(obra_raw)) if obra_raw else ''

        categoria_nome = ''
        categoria_raw = registro.get('categoria')
        if categoria_raw:
            try:
                conn = get_connection()
                cursor = conn.cursor(dictionary=True)
                cursor.execute("SELECT nome FROM categoria WHERE id = %s", (categoria_raw,))
                resultado = cursor.fetchone()
                cursor.close()
                conn.close()
                categoria_nome = resultado['nome'] if resultado else ''
            except:
                categoria_nome = ''

        # Status lançamento
        status = "Lançado" if registro.get('lancado') == 'Y' else "Pendente"

        row = [
            id_final,
            data_pagamento_final,
            valor_final,
            forma_pagamento_normalizada,
            quem_paga_normalizado,
            obra_normalizada,
            registro.get('titular', ''),
            registro.get('cpfCnpjTitularConta', ''),
            registro.get('chavePix', ''),
            registro.get('obra', ''),
            categoria_nome,
            status,
            registro.get('observacao', '')
        ]
        ws.append(row)

    # Formatar coluna de valor como número com ponto decimal (compatível com LibreOffice)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='right')
            # Garantir que é número e não texto
            if isinstance(cell.value, str):
                try:
                    cell.value = float(cell.value.lstrip("'"))
                except:
                    pass

    # Formatar coluna de data como data completa (dd/mm/yyyy)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = 'dd/mm/yyyy'
            # Garantir que é data e não texto
            if isinstance(cell.value, str):
                try:
                    from datetime import datetime as dt
                    cell.value = dt.strptime(cell.value.lstrip("'"), '%Y-%m-%d')
                except:
                    pass

    # Formatar coluna de ID como número (sem casas decimais)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
            if isinstance(cell.value, str):
                try:
                    cell.value = int(cell.value.lstrip("'"))
                except:
                    cell.value = str(cell.value).lstrip("'")

    # Salvar arquivo em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar arquivo como download
    return send_file(
        output,
        as_attachment=True,
        download_name=f"lancamentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )